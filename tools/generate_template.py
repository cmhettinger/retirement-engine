"""Generate the blank Retirement Engine workbook template."""

from __future__ import annotations

import re
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

ROOT_DIR = Path(__file__).resolve().parents[1]
WORKBOOK_RESOURCES_DIR = ROOT_DIR / "resources" / "workbooks"
TEMPLATE_PATH = WORKBOOK_RESOURCES_DIR / "template.xlsx"
EXAMPLE_PATH = WORKBOOK_RESOURCES_DIR / "example.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, size=12)
DEFAULT_FONT = Font(size=12)
DEFAULT_ROW_HEIGHT = 22
HEADER_ROW_HEIGHT = 24

CURRENCY_HEADERS = {
    "Annual",
    "Monthly",
    "Estimated Replacement Cost",
    "Current Balance",
    "Annual Contribution",
    "Employer Match",
    "Monthly Payment",
    "Annual Premium",
    "Coverage Amount",
}
PERCENT_HEADERS = {
    "Inflation Rate",
    "Expected Investment Return",
    "Estimated Tax Rate",
    "Interest Rate",
}
INTEGER_HEADERS = {
    "Current Age",
    "Planned Retirement Age",
    "Life Expectancy",
    "Social Security Claiming Age",
    "Planning Horizon Years",
    "Expected Useful Life",
    "Remaining Useful Life",
    "Next Replacement Year",
    "Payoff Year",
}


def yes_no_validation(field_name: str) -> DataValidation:
    validation = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
    )
    validation.error = "Select Yes, No, or leave blank."
    validation.errorTitle = f"Invalid {field_name} value"
    validation.prompt = "Select Yes or No, or leave blank if not classified yet."
    validation.promptTitle = field_name
    return validation


def slug(value: str) -> str:
    text = value.lower()
    text = text.replace("'s", "s")
    text = text.replace("401(k)", "401k")
    text = text.replace("403(b)", "403b")
    text = text.replace("457", "457")
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def person_key(value: str) -> str:
    return value.lower().replace(" ", "")


def append_rows(sheet: Worksheet, headers: list[str], rows: Iterable[list[object]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    format_sheet(sheet, headers)


def format_sheet(sheet: Worksheet, headers: list[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = DEFAULT_FONT

    sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT

    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    refresh_column_widths(sheet, headers)

    for index, header in enumerate(headers, start=1):
        if header in CURRENCY_HEADERS:
            number_format = '"$"#,##0'
        elif header in PERCENT_HEADERS:
            number_format = "0.0%"
        elif header in INTEGER_HEADERS:
            number_format = "0"
        else:
            number_format = None

        if number_format:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=index).number_format = number_format


def refresh_column_widths(sheet: Worksheet, headers: list[str]) -> None:
    for index, _header in enumerate(headers, start=1):
        column = get_column_letter(index)
        max_length = max(
            len(str(sheet.cell(row=row, column=index).value or ""))
            for row in range(1, sheet.max_row + 1)
        )
        sheet.column_dimensions[column].width = min(max(max_length + 2, 12), 38)


def format_assumption_values(sheet: Worksheet) -> None:
    for row in range(2, sheet.max_row + 1):
        assumption = str(sheet.cell(row=row, column=2).value or "")
        value_cell = sheet.cell(row=row, column=3)
        if assumption in PERCENT_HEADERS:
            value_cell.number_format = "0.0%"
        elif assumption in INTEGER_HEADERS:
            value_cell.number_format = "0"
        elif assumption in {"Desired Cash Reserve", "Desired Estate Value"}:
            value_cell.number_format = '"$"#,##0'


def add_yes_no_validation(sheet: Worksheet, column: str, field_name: str) -> None:
    validation = yes_no_validation(field_name)
    sheet.add_data_validation(validation)
    validation.add(f"{column}2:{column}{sheet.max_row}")


def build_assumptions() -> tuple[list[str], list[list[object]]]:
    headers = ["Person / Scope", "Assumption", "Value", "Notes", "ID"]
    rows: list[list[object]] = [
        ["System", "Workbook Version", "0.1.0", "", "system.workbook.version"]
    ]

    for person in ("Person 1", "Person 2"):
        key = person_key(person)
        rows.append([person, "Name", "", "", f"assumptions.{key}.name"])
        for assumption in (
            "Current Age",
            "Planned Retirement Age",
            "Life Expectancy",
            "Social Security Claiming Age",
        ):
            rows.append([person, assumption, "", "", f"assumptions.{key}.{slug(assumption)}"])

    for assumption in (
        "Planning Horizon Years",
        "Inflation Rate",
        "Expected Investment Return",
        "Estimated Tax Rate",
        "Desired Cash Reserve",
        "Desired Estate Value",
    ):
        rows.append(["Household", assumption, "", "", f"assumptions.household.{slug(assumption)}"])

    return headers, rows


def budget_rows() -> tuple[list[str], list[list[object]]]:
    headers = ["Must Pay", "Category", "Item", "Annual", "Monthly", "Notes", "ID"]
    categories = {
        "Housing": [
            "Mortgage / Rent",
            "Property Taxes",
            "Homeowners / Renters Insurance",
            "HOA / Condo Fees",
            "Home Maintenance",
            "Lawn Care",
            "Snow Removal",
            "Pest Control",
            "House Cleaning",
            "Handyman Services",
            "Window / Gutter Cleaning",
            "Tree Removal",
            "Household Supplies",
        ],
        "Utilities": [
            "Electricity",
            "Natural Gas / Propane / Heating Oil",
            "Water",
            "Sewer",
            "Trash / Recycling",
            "Generator Fuel / Backup Power",
        ],
        "Food": ["Groceries", "Restaurants", "Coffee / Snacks", "Alcohol", "Meal Delivery"],
        "Healthcare": [
            "Health Insurance Premiums",
            "Medicare Premiums",
            "Medicare Supplement",
            "Dental Insurance",
            "Vision Insurance",
            "Prescription Medications",
            "Over-the-Counter Medications",
            "Doctor Visits",
            "Specialists",
            "Dental Care",
            "Vision Care",
            "Hearing Care",
            "Medical Equipment",
            "Long-Term Care Insurance",
            "Out-of-Pocket Medical",
            "Gym / Wellness",
        ],
        "Insurance": [
            "Life Insurance",
            "Umbrella Liability",
            "Identity Theft Protection",
            "Pet Insurance",
            "Flood Insurance",
            "Earthquake Insurance",
            "RV / Boat / Motorcycle Insurance",
        ],
        "Taxes": [
            "Federal Income Tax",
            "State Income Tax",
            "Local Taxes",
            "Capital Gains Tax",
            "Estimated Tax Payments",
        ],
        "Transportation": [
            "Vehicle Payments",
            "Fuel",
            "Auto Insurance",
            "Registration / Inspection",
            "Driver's License Renewal",
            "Vehicle Property Tax",
            "Maintenance",
            "Tires",
            "Repairs",
            "Parking",
            "Tolls",
            "Public Transportation",
            "Ride Sharing",
        ],
        "Communications": [
            "Internet",
            "Mobile Phones",
            "Home Phone",
            "Cable / Satellite TV",
            "Streaming Services",
            "Cloud Storage",
        ],
        "Personal": [
            "Clothing",
            "Shoes",
            "Haircuts",
            "Personal Care",
            "Toiletries",
            "Laundry / Dry Cleaning",
        ],
        "Family": [
            "Gifts",
            "Holidays",
            "Grandchildren",
            "Charitable Giving",
            "Family Support",
            "Weddings / Funerals",
        ],
        "Pets": ["Pet Food", "Veterinary Care", "Grooming", "Boarding", "Medications"],
        "Financial": [
            "Investment Management Fees",
            "Financial Advisor",
            "Tax Preparation",
            "Tax Software",
            "Brokerage / Custodial Fees",
            "Banking Fees",
            "Safe Deposit Box",
        ],
        "Debt": [
            "Credit Cards",
            "Personal Loans",
            "Home Equity Loan",
            "Student Loans",
            "Other Debt",
        ],
        "Legal": [
            "Estate Planning",
            "Attorney / Legal Services",
            "Trust & Will Updates",
            "Notary / Filing Fees",
            "Document Storage",
        ],
        "Memberships": [
            "Gym Membership",
            "Warehouse Clubs",
            "Professional Organizations",
            "Clubs / Associations",
            "Software Subscriptions",
        ],
        "Recreation": [
            "Travel",
            "Vacations",
            "Hotels",
            "Cruises",
            "Hobbies",
            "Sporting Events",
            "Concerts",
            "Clubs",
            "Golf",
            "Fishing / Hunting",
            "Books",
            "Movies",
            "Entertainment Subscriptions",
            "Education / Classes",
            "National / State Park Passes",
        ],
        "Replacement Reserve": ["Annual Replacement Reserve Contribution"],
        "Other": [
            "Amazon / General Shopping",
            "Office Supplies",
            "Postage / Shipping",
            "Miscellaneous",
            "Contingency",
        ],
    }
    rows = [
        ["", category, item, "", "", "", f"budget.{slug(category)}.{slug(item)}"]
        for category, items in categories.items()
        for item in items
    ]
    return headers, rows


def reserve_rows() -> tuple[list[str], list[list[object]]]:
    headers = [
        "Reserve Item",
        "Estimated Replacement Cost",
        "Expected Useful Life",
        "Current Age",
        "Remaining Useful Life",
        "Next Replacement Year",
        "Notes",
        "ID",
    ]
    items = [
        "Emergency Fund Replenishment",
        "Roof",
        "HVAC System",
        "Water Heater",
        "Plumbing Repairs",
        "Electrical Repairs",
        "Appliance Replacement",
        "Flooring / Carpet",
        "Exterior Painting",
        "Remodeling / Renovation",
        "Windows / Doors",
        "Driveway / Walkway",
        "Deck / Patio",
        "Landscaping",
        "Furniture",
        "Mattress Replacement",
        "Vehicle Replacement",
        "Vehicle Major Repairs",
        "Computer Replacement",
        "Phone Replacement",
        "Tablet Replacement",
        "Television / Electronics",
        "Camera Equipment",
        "Travel Fund",
        "Medical Contingency",
        "Long-Term Care Reserve",
        "Family Assistance Reserve",
        "Pet Medical Reserve",
        "Other",
    ]
    rows = [[item, "", "", "", "", "", "", f"reserve.{slug(item)}"] for item in items]
    return headers, rows


def income_rows() -> tuple[list[str], list[list[object]]]:
    headers = ["Owner", "Income Source", "Annual", "Monthly", "Taxable", "Notes", "ID"]
    items = [
        ("Person 1", "Social Security"),
        ("Person 2", "Social Security"),
        ("Person 1", "Pension"),
        ("Person 2", "Pension"),
        ("Household", "Traditional IRA Withdrawals"),
        ("Household", "Roth IRA Withdrawals"),
        ("Household", "401(k) Withdrawals"),
        ("Household", "403(b) / 457 Withdrawals"),
        ("Household", "Taxable Brokerage Withdrawals"),
        ("Household", "Dividends"),
        ("Household", "Interest Income"),
        ("Household", "Rental Income"),
        ("Household", "Business Income"),
        ("Household", "Annuities"),
        ("Household", "Part-Time Employment"),
        ("Household", "VA Benefits"),
        ("Household", "Disability Benefits"),
        ("Household", "Other Income"),
    ]
    rows = [
        [owner, source, "", "", "", "", f"income.{person_key(owner)}.{slug(source)}"]
        for owner, source in items
    ]
    return headers, rows


def asset_rows() -> tuple[list[str], list[list[object]]]:
    headers = [
        "Owner",
        "Account Type",
        "Institution",
        "Current Balance",
        "Annual Contribution",
        "Employer Match",
        "Tax Treatment",
        "Notes",
        "ID",
    ]
    items = [
        ("Person 1", "401(k)"),
        ("Person 1", "Traditional IRA"),
        ("Person 1", "Roth IRA"),
        ("Person 1", "HSA"),
        ("Person 2", "401(k)"),
        ("Person 2", "Traditional IRA"),
        ("Person 2", "Roth IRA"),
        ("Person 2", "HSA"),
        ("Household", "Taxable Brokerage"),
        ("Household", "Cash / Savings"),
        ("Household", "CDs / Money Market"),
        ("Household", "Real Estate"),
        ("Household", "Other Assets"),
    ]
    rows = [
        [owner, account, "", "", "", "", "", "", f"asset.{person_key(owner)}.{slug(account)}"]
        for owner, account in items
    ]
    return headers, rows


def liability_rows() -> tuple[list[str], list[list[object]]]:
    headers = [
        "Owner",
        "Liability Type",
        "Lender",
        "Current Balance",
        "Interest Rate",
        "Monthly Payment",
        "Payoff Year",
        "Notes",
        "ID",
    ]
    items = [
        ("Household", "Mortgage"),
        ("Household", "Home Equity Loan"),
        ("Household", "Vehicle Loan"),
        ("Household", "Credit Card Debt"),
        ("Household", "Personal Loan"),
        ("Household", "Student Loan"),
        ("Household", "Medical Debt"),
        ("Household", "Other Liability"),
    ]
    rows = [
        [
            owner,
            liability,
            "",
            "",
            "",
            "",
            "",
            "",
            f"liability.{person_key(owner)}.{slug(liability)}",
        ]
        for owner, liability in items
    ]
    return headers, rows


def insurance_rows() -> tuple[list[str], list[list[object]]]:
    headers = [
        "Owner",
        "Policy Type",
        "Provider",
        "Annual Premium",
        "Coverage Amount",
        "Beneficiary",
        "Notes",
        "ID",
    ]
    items = [
        ("Person 1", "Life Insurance"),
        ("Person 2", "Life Insurance"),
        ("Person 1", "Disability Insurance"),
        ("Person 2", "Disability Insurance"),
        ("Person 1", "Long-Term Care Insurance"),
        ("Person 2", "Long-Term Care Insurance"),
        ("Household", "Health Insurance"),
        ("Household", "Dental Insurance"),
        ("Household", "Vision Insurance"),
        ("Household", "Homeowners Insurance"),
        ("Household", "Auto Insurance"),
        ("Household", "Umbrella Insurance"),
    ]
    rows = [
        [owner, policy, "", "", "", "", "", f"insurance.{person_key(owner)}.{slug(policy)}"]
        for owner, policy in items
    ]
    return headers, rows


def build_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = [
        ("Assumptions", build_assumptions()),
        ("Budget", budget_rows()),
        ("Reserves", reserve_rows()),
        ("Income", income_rows()),
        ("Assets", asset_rows()),
        ("Liabilities", liability_rows()),
        ("Insurance", insurance_rows()),
    ]

    for name, (headers, rows) in sheets:
        sheet = workbook.create_sheet(name)
        append_rows(sheet, headers, rows)
        if name == "Assumptions":
            format_assumption_values(sheet)
        elif name == "Budget":
            add_yes_no_validation(sheet, "A", "Must Pay")
        elif name == "Income":
            add_yes_no_validation(sheet, "E", "Taxable")

    workbook.active = 0
    return workbook


def headers_for(sheet: Worksheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in sheet[1]}


def rows_by_id(sheet: Worksheet) -> dict[str, int]:
    return {
        str(sheet.cell(row=row, column=sheet.max_column).value): row
        for row in range(2, sheet.max_row + 1)
    }


def set_values_by_id(
    sheet: Worksheet,
    row_id: str,
    values: dict[str, object],
) -> None:
    headers = headers_for(sheet)
    row = rows_by_id(sheet)[row_id]
    for header, value in values.items():
        sheet.cell(row=row, column=headers[header]).value = value


def set_budget(
    sheet: Worksheet,
    row_id: str,
    must_pay: str,
    annual: int | None = None,
    monthly: int | None = None,
    notes: str = "",
) -> None:
    values: dict[str, object] = {"Must Pay": must_pay, "Notes": notes}
    if annual is not None:
        values["Annual"] = annual
    if monthly is not None:
        values["Monthly"] = monthly
    set_values_by_id(sheet, row_id, values)


def set_reserve(
    sheet: Worksheet,
    row_id: str,
    cost: int,
    useful_life: int,
    current_age: int,
    remaining_life: int,
    next_year: int,
    notes: str = "",
) -> None:
    set_values_by_id(
        sheet,
        row_id,
        {
            "Estimated Replacement Cost": cost,
            "Expected Useful Life": useful_life,
            "Current Age": current_age,
            "Remaining Useful Life": remaining_life,
            "Next Replacement Year": next_year,
            "Notes": notes,
        },
    )


def set_income(
    sheet: Worksheet,
    row_id: str,
    taxable: str,
    annual: int | None = None,
    monthly: int | None = None,
    notes: str = "",
) -> None:
    values: dict[str, object] = {"Taxable": taxable, "Notes": notes}
    if annual is not None:
        values["Annual"] = annual
    if monthly is not None:
        values["Monthly"] = monthly
    set_values_by_id(sheet, row_id, values)


def set_asset(
    sheet: Worksheet,
    row_id: str,
    institution: str,
    balance: int,
    contribution: int,
    employer_match: int,
    tax_treatment: str,
    notes: str = "",
) -> None:
    set_values_by_id(
        sheet,
        row_id,
        {
            "Institution": institution,
            "Current Balance": balance,
            "Annual Contribution": contribution,
            "Employer Match": employer_match,
            "Tax Treatment": tax_treatment,
            "Notes": notes,
        },
    )


def set_liability(
    sheet: Worksheet,
    row_id: str,
    lender: str,
    balance: int,
    rate: float,
    monthly_payment: int,
    payoff_year: int,
    notes: str = "",
) -> None:
    set_values_by_id(
        sheet,
        row_id,
        {
            "Lender": lender,
            "Current Balance": balance,
            "Interest Rate": rate,
            "Monthly Payment": monthly_payment,
            "Payoff Year": payoff_year,
            "Notes": notes,
        },
    )


def set_insurance(
    sheet: Worksheet,
    row_id: str,
    provider: str,
    annual_premium: int,
    coverage_amount: int,
    beneficiary: str,
    notes: str = "",
) -> None:
    set_values_by_id(
        sheet,
        row_id,
        {
            "Provider": provider,
            "Annual Premium": annual_premium,
            "Coverage Amount": coverage_amount,
            "Beneficiary": beneficiary,
            "Notes": notes,
        },
    )


def populate_example_workbook(workbook: Workbook) -> None:
    populate_example_assumptions(workbook["Assumptions"])
    populate_example_budget(workbook["Budget"])
    populate_example_reserves(workbook["Reserves"])
    populate_example_income(workbook["Income"])
    populate_example_assets(workbook["Assets"])
    populate_example_liabilities(workbook["Liabilities"])
    populate_example_insurance(workbook["Insurance"])

    for sheet in workbook.worksheets:
        refresh_column_widths(sheet, [str(cell.value) for cell in sheet[1]])


def populate_example_assumptions(sheet: Worksheet) -> None:
    values = {
        "system.workbook.version": "0.1.0",
        "assumptions.person1.name": "Han Solo",
        "assumptions.person1.current_age": 58,
        "assumptions.person1.planned_retirement_age": 67,
        "assumptions.person1.life_expectancy": 90,
        "assumptions.person1.social_security_claiming_age": 70,
        "assumptions.person2.name": "Leia Organa",
        "assumptions.person2.current_age": 56,
        "assumptions.person2.planned_retirement_age": 65,
        "assumptions.person2.life_expectancy": 92,
        "assumptions.person2.social_security_claiming_age": 67,
        "assumptions.household.planning_horizon_years": 35,
        "assumptions.household.inflation_rate": 0.03,
        "assumptions.household.expected_investment_return": 0.055,
        "assumptions.household.estimated_tax_rate": 0.18,
        "assumptions.household.desired_cash_reserve": 45000,
        "assumptions.household.desired_estate_value": 250000,
    }
    for row_id, value in values.items():
        set_values_by_id(sheet, row_id, {"Value": value})


def populate_example_budget(sheet: Worksheet) -> None:
    entries = [
        ("budget.housing.mortgage_rent", "Yes", None, 2450, "Primary residence mortgage"),
        ("budget.housing.property_taxes", "Yes", 7200, None, "Annual county tax bill"),
        ("budget.housing.homeowners_renters_insurance", "Yes", 1800, None, ""),
        ("budget.housing.hoa_condo_fees", "No", None, 125, ""),
        ("budget.housing.home_maintenance", "Yes", 4200, None, ""),
        ("budget.housing.lawn_care", "No", None, 150, "Seasonal average"),
        ("budget.housing.snow_removal", "No", 600, None, ""),
        ("budget.housing.house_cleaning", "No", None, 220, ""),
        ("budget.housing.household_supplies", "Yes", None, 140, ""),
        ("budget.utilities.electricity", "Yes", None, 185, ""),
        ("budget.utilities.natural_gas_propane_heating_oil", "Yes", None, 110, ""),
        ("budget.utilities.water", "Yes", None, 65, ""),
        ("budget.utilities.sewer", "Yes", None, 45, ""),
        ("budget.utilities.trash_recycling", "Yes", None, 35, ""),
        ("budget.food.groceries", "Yes", None, 850, ""),
        ("budget.food.restaurants", "No", None, 325, ""),
        ("budget.food.coffee_snacks", "No", None, 80, ""),
        ("budget.food.alcohol", "No", None, 60, ""),
        ("budget.healthcare.health_insurance_premiums", "Yes", None, 650, ""),
        ("budget.healthcare.dental_insurance", "Yes", None, 55, ""),
        ("budget.healthcare.vision_insurance", "Yes", None, 25, ""),
        ("budget.healthcare.prescription_medications", "Yes", None, 120, ""),
        ("budget.healthcare.doctor_visits", "Yes", 900, None, ""),
        ("budget.healthcare.dental_care", "Yes", 1000, None, ""),
        ("budget.healthcare.out_of_pocket_medical", "Yes", 2400, None, ""),
        ("budget.healthcare.gym_wellness", "No", None, 95, ""),
        ("budget.insurance.life_insurance", "Yes", 1200, None, ""),
        ("budget.insurance.umbrella_liability", "Yes", 420, None, ""),
        ("budget.insurance.identity_theft_protection", "No", None, 18, ""),
        ("budget.insurance.pet_insurance", "No", None, 55, ""),
        ("budget.taxes.federal_income_tax", "Yes", 18500, None, ""),
        ("budget.taxes.state_income_tax", "Yes", 6200, None, ""),
        ("budget.transportation.vehicle_payments", "Yes", None, 475, ""),
        ("budget.transportation.fuel", "Yes", None, 260, ""),
        ("budget.transportation.auto_insurance", "Yes", 2100, None, ""),
        ("budget.transportation.registration_inspection", "Yes", 260, None, ""),
        ("budget.transportation.maintenance", "Yes", 1200, None, ""),
        ("budget.transportation.tires", "Yes", 700, None, ""),
        ("budget.transportation.repairs", "Yes", 900, None, ""),
        ("budget.transportation.parking", "No", None, 40, ""),
        ("budget.transportation.tolls", "No", None, 35, ""),
        ("budget.communications.internet", "Yes", None, 90, ""),
        ("budget.communications.mobile_phones", "Yes", None, 165, ""),
        ("budget.communications.streaming_services", "No", None, 75, ""),
        ("budget.communications.cloud_storage", "No", 120, None, ""),
        ("budget.personal.clothing", "No", 1800, None, ""),
        ("budget.personal.haircuts", "No", None, 90, ""),
        ("budget.personal.personal_care", "No", None, 70, ""),
        ("budget.personal.toiletries", "Yes", None, 65, ""),
        ("budget.family.gifts", "No", 1600, None, ""),
        ("budget.family.holidays", "No", 2200, None, ""),
        ("budget.family.charitable_giving", "No", 2400, None, ""),
        ("budget.family.family_support", "No", 1800, None, ""),
        ("budget.pets.pet_food", "Yes", None, 85, ""),
        ("budget.pets.veterinary_care", "Yes", 900, None, ""),
        ("budget.pets.grooming", "No", None, 55, ""),
        ("budget.financial.investment_management_fees", "No", 1800, None, ""),
        ("budget.financial.tax_preparation", "Yes", 650, None, ""),
        ("budget.financial.banking_fees", "No", None, 15, ""),
        ("budget.debt.credit_cards", "Yes", None, 250, "Temporary payoff payment"),
        ("budget.debt.personal_loans", "No", None, 0, ""),
        ("budget.legal.estate_planning", "No", 1000, None, "Periodic updates"),
        ("budget.memberships.warehouse_clubs", "No", 120, None, ""),
        ("budget.memberships.software_subscriptions", "No", None, 45, ""),
        ("budget.recreation.travel", "No", 4800, None, ""),
        ("budget.recreation.vacations", "No", 3500, None, ""),
        ("budget.recreation.hobbies", "No", None, 180, ""),
        ("budget.recreation.books", "No", None, 40, ""),
        ("budget.recreation.movies", "No", None, 45, ""),
        ("budget.recreation.education_classes", "No", 600, None, ""),
        (
            "budget.replacement_reserve.annual_replacement_reserve_contribution",
            "Yes",
            7500,
            None,
            "",
        ),
        ("budget.other.amazon_general_shopping", "No", None, 250, ""),
        ("budget.other.office_supplies", "No", 240, None, ""),
        ("budget.other.miscellaneous", "No", None, 150, ""),
        ("budget.other.contingency", "Yes", 3000, None, ""),
    ]
    for row_id, must_pay, annual, monthly, notes in entries:
        set_budget(sheet, row_id, must_pay, annual, monthly, notes)


def populate_example_reserves(sheet: Worksheet) -> None:
    entries = [
        ("reserve.emergency_fund_replenishment", 45000, 5, 2, 3, 2029, "Target reserve rebuild"),
        ("reserve.roof", 26000, 25, 14, 11, 2037, "Architectural shingles"),
        ("reserve.hvac_system", 14000, 18, 9, 9, 2035, ""),
        ("reserve.water_heater", 2200, 12, 7, 5, 2031, ""),
        ("reserve.plumbing_repairs", 5000, 15, 5, 10, 2036, ""),
        ("reserve.electrical_repairs", 4500, 20, 8, 12, 2038, ""),
        ("reserve.appliance_replacement", 8500, 10, 4, 6, 2032, ""),
        ("reserve.flooring_carpet", 12000, 15, 6, 9, 2035, ""),
        ("reserve.exterior_painting", 6500, 10, 3, 7, 2033, ""),
        ("reserve.furniture", 9000, 12, 5, 7, 2033, ""),
        ("reserve.mattress_replacement", 2500, 10, 6, 4, 2030, ""),
        ("reserve.vehicle_replacement", 42000, 10, 5, 5, 2031, "Replace primary vehicle"),
        ("reserve.vehicle_major_repairs", 4500, 5, 2, 3, 2029, ""),
        ("reserve.computer_replacement", 3000, 5, 3, 2, 2028, ""),
        ("reserve.phone_replacement", 1800, 4, 2, 2, 2028, "Two phones"),
        ("reserve.tablet_replacement", 900, 5, 1, 4, 2030, ""),
        ("reserve.television_electronics", 2500, 8, 3, 5, 2031, ""),
        ("reserve.travel_fund", 6000, 1, 0, 1, 2027, "Annual travel reserve"),
        ("reserve.medical_contingency", 15000, 5, 1, 4, 2030, ""),
        ("reserve.long_term_care_reserve", 30000, 10, 0, 10, 2036, ""),
        ("reserve.family_assistance_reserve", 10000, 5, 1, 4, 2030, ""),
        ("reserve.pet_medical_reserve", 3500, 4, 1, 3, 2029, ""),
    ]
    for entry in entries:
        set_reserve(sheet, *entry)


def populate_example_income(sheet: Worksheet) -> None:
    entries = [
        ("income.person1.social_security", "Yes", 39000, None, "Estimated annual benefit at 70"),
        ("income.person2.social_security", "Yes", 32000, None, "Estimated annual benefit at 67"),
        ("income.person1.pension", "Yes", 12000, None, "Small legacy pension"),
        (
            "income.household.traditional_ira_withdrawals",
            "Yes",
            36000,
            None,
            "Planned retirement draw",
        ),
        ("income.household.roth_ira_withdrawals", "No", 12000, None, "Tax-free supplemental draw"),
        ("income.household.401k_withdrawals", "Yes", 24000, None, ""),
        ("income.household.taxable_brokerage_withdrawals", "Yes", 18000, None, ""),
        ("income.household.dividends", "Yes", None, 650, "Taxable brokerage dividends"),
        ("income.household.interest_income", "Yes", None, 275, "Cash and CD interest"),
        ("income.household.rental_income", "Yes", None, 0, ""),
        ("income.household.part_time_employment", "Yes", 10000, None, "Optional consulting"),
    ]
    for row_id, taxable, annual, monthly, notes in entries:
        set_income(sheet, row_id, taxable, annual, monthly, notes)


def populate_example_assets(sheet: Worksheet) -> None:
    entries = [
        (
            "asset.person1.401k",
            "Corellian Manufacturing 401(k)",
            425000,
            19500,
            6500,
            "Pre-tax",
            "",
        ),
        ("asset.person1.traditional_ira", "Vanguard", 185000, 0, 0, "Pre-tax", "Rollover IRA"),
        ("asset.person1.roth_ira", "Fidelity", 92000, 7500, 0, "Roth", ""),
        ("asset.person1.hsa", "HealthEquity", 36000, 4150, 0, "HSA", ""),
        (
            "asset.person2.401k",
            "Alderaan Public Service 403(b)",
            390000,
            18000,
            5500,
            "Pre-tax",
            "",
        ),
        ("asset.person2.traditional_ira", "Schwab", 165000, 0, 0, "Pre-tax", ""),
        ("asset.person2.roth_ira", "Fidelity", 110000, 7500, 0, "Roth", ""),
        ("asset.person2.hsa", "Optum", 28000, 4150, 0, "HSA", ""),
        ("asset.household.taxable_brokerage", "Vanguard", 275000, 12000, 0, "Taxable", ""),
        (
            "asset.household.cash_savings",
            "Local Credit Union",
            52000,
            0,
            0,
            "Cash",
            "Emergency reserve",
        ),
        ("asset.household.cds_money_market", "Ally Bank", 85000, 0, 0, "Cash", "Laddered CDs"),
        (
            "asset.household.real_estate",
            "Primary residence",
            520000,
            0,
            0,
            "Real Estate",
            "Estimated market value",
        ),
        (
            "asset.household.other_assets",
            "Household vehicles",
            38000,
            0,
            0,
            "Personal Property",
            "",
        ),
    ]
    for entry in entries:
        set_asset(sheet, *entry)


def populate_example_liabilities(sheet: Worksheet) -> None:
    entries = [
        (
            "liability.household.mortgage",
            "First Republic Credit Union",
            238000,
            0.041,
            1850,
            2042,
            "",
        ),
        ("liability.household.vehicle_loan", "Toyota Financial", 18000, 0.049, 475, 2029, ""),
        (
            "liability.household.credit_card_debt",
            "Chase",
            3200,
            0.189,
            250,
            2027,
            "Temporary balance",
        ),
        (
            "liability.household.other_liability",
            "Furniture store promotion",
            2100,
            0.0,
            175,
            2027,
            "",
        ),
    ]
    for entry in entries:
        set_liability(sheet, *entry)


def populate_example_insurance(sheet: Worksheet) -> None:
    entries = [
        (
            "insurance.person1.life_insurance",
            "Northwestern Mutual",
            780,
            250000,
            "Leia Organa",
            "20-year term",
        ),
        (
            "insurance.person2.life_insurance",
            "Northwestern Mutual",
            720,
            250000,
            "Han Solo",
            "20-year term",
        ),
        ("insurance.person1.long_term_care_insurance", "Genworth", 2100, 165000, "Han Solo", ""),
        ("insurance.person2.long_term_care_insurance", "Genworth", 2300, 165000, "Leia Organa", ""),
        (
            "insurance.household.health_insurance",
            "Blue Cross",
            7800,
            0,
            "Household",
            "Pre-Medicare estimate",
        ),
        ("insurance.household.dental_insurance", "Delta Dental", 660, 2000, "Household", ""),
        ("insurance.household.vision_insurance", "VSP", 300, 300, "Household", ""),
        ("insurance.household.homeowners_insurance", "State Farm", 1800, 520000, "Household", ""),
        (
            "insurance.household.auto_insurance",
            "State Farm",
            2100,
            300000,
            "Household",
            "Two vehicles",
        ),
        ("insurance.household.umbrella_insurance", "State Farm", 420, 1000000, "Household", ""),
    ]
    for entry in entries:
        set_insurance(sheet, *entry)


def main() -> None:
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)

    workbook = build_workbook()
    workbook.save(TEMPLATE_PATH)
    print(f"Generated {TEMPLATE_PATH}")

    example_workbook = build_workbook()
    populate_example_workbook(example_workbook)
    example_workbook.save(EXAMPLE_PATH)
    print(f"Generated {EXAMPLE_PATH}")


if __name__ == "__main__":
    main()
