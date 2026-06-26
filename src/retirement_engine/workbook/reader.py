"""Workbook loading for Retirement Engine spreadsheets."""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import cast

from openpyxl import load_workbook

type WorkbookCellValue = str | int | float | bool | date | datetime | time | timedelta | None


@dataclass(frozen=True)
class WorkbookRow:
    """A parsed worksheet row keyed by the sheet's header row."""

    row_number: int
    values: Mapping[str, WorkbookCellValue]

    @property
    def row_id(self) -> str | None:
        value = self.values.get("ID")
        if value is None:
            return None
        return str(value)


@dataclass(frozen=True)
class WorkbookSheet:
    """A parsed worksheet with headers and data rows."""

    name: str
    headers: tuple[str, ...]
    rows: tuple[WorkbookRow, ...]

    @property
    def rows_by_id(self) -> Mapping[str, WorkbookRow]:
        return {row.row_id: row for row in self.rows if row.row_id is not None}


@dataclass(frozen=True)
class RetirementWorkbook:
    """Structured representation of a Retirement Engine workbook."""

    path: Path
    sheets: Mapping[str, WorkbookSheet]

    def sheet(self, name: str) -> WorkbookSheet:
        return self.sheets[name]


def load_retirement_workbook(path: str | Path) -> RetirementWorkbook:
    """Load workbook sheets into simple Python structures without calculations."""

    workbook_path = Path(path).expanduser().resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    try:
        sheets = {
            worksheet.title: _parse_sheet(
                worksheet.title,
                tuple(worksheet.iter_rows(values_only=True)),
            )
            for worksheet in workbook.worksheets
        }
    finally:
        workbook.close()

    return RetirementWorkbook(path=workbook_path, sheets=sheets)


def _parse_sheet(
    name: str,
    rows: tuple[tuple[object, ...], ...],
) -> WorkbookSheet:
    if not rows:
        return WorkbookSheet(name=name, headers=(), rows=())

    headers = tuple("" if value is None else str(value).strip() for value in rows[0])
    parsed_rows = tuple(
        WorkbookRow(row_number=row_number, values=_row_values(headers, row))
        for row_number, row in enumerate(rows[1:], start=2)
        if _has_value(row)
    )

    return WorkbookSheet(name=name, headers=headers, rows=parsed_rows)


def _row_values(
    headers: tuple[str, ...],
    row: tuple[object, ...],
) -> Mapping[str, WorkbookCellValue]:
    return {
        header: _cell_value(row[index]) if index < len(row) else None
        for index, header in enumerate(headers)
    }


def _cell_value(value: object) -> WorkbookCellValue:
    return cast(WorkbookCellValue, value)


def _has_value(row: tuple[object, ...]) -> bool:
    return any(value is not None for value in row)
